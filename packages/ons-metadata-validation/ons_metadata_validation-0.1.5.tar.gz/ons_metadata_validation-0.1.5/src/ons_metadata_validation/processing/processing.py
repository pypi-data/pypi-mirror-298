import logging
from typing import Dict, List, Tuple

import attr
import pandas as pd
from attr.validators import instance_of
from openpyxl import Workbook, load_workbook
from tqdm import tqdm

from ons_metadata_validation.fixing.fixing_utils import fix_main
from ons_metadata_validation.io.cell_template import MetadataValues
from ons_metadata_validation.io.input_functions import (
    apply_version_changes,
    check_metadata_vs_map,
    extract_all_metadata_values,
    load_metadata_cells,
)
from ons_metadata_validation.processing.developing import (
    md_values_to_df,
    search_md_values,
)
from ons_metadata_validation.processing.outputs_adapter import (
    get_df_all_fails,
    get_tab_lens,
)
from ons_metadata_validation.reference.delta_table import DELTA_TABLE
from ons_metadata_validation.reference.role_based_configs import (
    ROLE_CONFIGS,
)
from ons_metadata_validation.reference.v2_template import V2_TEMPLATE
from ons_metadata_validation.utils.logger import (
    compress_logging_value,
)
from ons_metadata_validation.validation import (
    comparative_validations as comp_val,
)

logger = logging.getLogger()


@attr.define
class MetadataProcessor:
    md_filepath: str = attr.ib(validator=[instance_of(str)])
    variable_check_set: str = attr.ib(validator=[instance_of(str)])
    save_report: bool = attr.ib(validator=[instance_of(bool)])
    save_corrected_copy: bool = attr.ib(validator=[instance_of(bool)])
    target_version: str = attr.ib(validator=[instance_of(str)], init=False)
    template_map: dict = attr.ib(validator=[instance_of(dict)], init=False)
    fails_dict: dict = attr.ib(validator=[instance_of(dict)], init=False)
    tab_lens: dict = attr.ib(validator=[instance_of(dict)], init=False)
    fails_df: pd.DataFrame = attr.ib(
        validator=[instance_of(pd.DataFrame)], init=False
    )

    def main_process(self) -> None:
        """the main process for the validation

        Args:
            md_filepath (str): the filepath to the metadata template
            save_report (bool, optional): Whether to save the report. Defaults to True.
            save_corrected_copy (bool, optional): Whether to save a corrected version of the template. Defaults to False.

        Raises:
            ValueError: if metadata map and input template don't match
            RuntimeError: if fails to remove all the null rows
            RuntimeError: if fails to validate all metadata values
            RuntimeError: if fails to save to excel

        Returns:
            Dict[str, pd.DataFrame]: the fails dict of hard and soft fails
        """
        for key, val in locals().items():
            logger.debug(f"{key} = {compress_logging_value(val)}")

        wb = load_workbook(self.md_filepath, data_only=True)
        self.target_version = wb["Dataset Resource"]["A1"].value
        self.template_map = apply_version_changes(
            V2_TEMPLATE, DELTA_TABLE, self.target_version
        )
        md_refs = load_metadata_cells(self.template_map)

        role_config = ROLE_CONFIGS[self.variable_check_set.lower()]

        non_matches: Dict[str, str] = check_metadata_vs_map(
            self.md_filepath, md_refs
        )
        if non_matches:
            logger.error(
                f"metadata file does not match map {non_matches}. "
                "Check if using the correct metadata map or make new one"
            )
            raise ValueError(
                f"metadata file does not match map {non_matches}. "
                "Check if using the correct metadata map or make new one"
            )

        metadata_values = extract_all_metadata_values(wb, md_refs)

        if not _remove_null_rows_for_tabs(
            metadata_values,
            [
                "Dataset File",
                "Dataset Series",
                "Variables",
                "Codes and Values",
            ],
        ):
            logger.error(
                "_remove_null_rows_for_tabs() failed to remove all null rows"
            )
            raise RuntimeError(
                "_remove_null_rows_for_tabs() failed to remove all null rows"
            )

        self.tab_lens = get_tab_lens(metadata_values)

        if not _validate_metadata_values(metadata_values, role_config):
            logger.error(
                "_validate_metadata_values() failed to validate all metadata values"
            )
            raise RuntimeError(
                "_validate_metadata_values() failed to validate all metadata values"
            )

        self.fails_dict = _get_all_fails_dicts(metadata_values)
        self.fails_dict["comparative_fails"] = _get_comparative_fails(
            metadata_values, self.template_map
        )

        for fail_type, fail_list in tqdm(
            self.fails_dict.items(), "finding hard and soft fails"
        ):
            self.fails_dict[fail_type] = _add_cell_references_to_df(
                wb, metadata_values, fail_list
            )

        if self.save_corrected_copy:
            fix_main(wb, self.fails_dict, self.md_filepath)

        self.fails_df = get_df_all_fails(self.fails_dict)


def _remove_null_rows_for_tabs(
    metadata_values: Dict[str, MetadataValues], tabs: List[str]
) -> bool:
    """remove rows for all tabs that are all(null or have a formula in them).
    Wrapper function for _remove_null_rows()

    Args:
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        tabs (List[str]): list of tabs to remove nulls from

    Returns:
        bool: True if all are successful
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    successes = []

    for tab in tabs:
        successes.append(_remove_null_rows(metadata_values, tab))

    return all(successes)


def _remove_null_rows(
    metadata_values: Dict[str, MetadataValues], tab: str
) -> bool:
    """remove rows for given tab that are all(null or have a formula in them).

    Args:
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        tab (str): tab to remove nulls from

    Returns:
        bool: True if successful
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    try:
        tab_mvs = _get_md_values_by_tab(metadata_values, tab)
        tab_values = [mv.values for mv in tab_mvs]
        tab_records = list(zip(*tab_values))

        null_idx = []

        for i, record in tqdm(
            enumerate(tab_records), f"finding null rows for {tab} tab"
        ):
            # all the Nones are cast to str, and formulas are dragged down to a certain
            # row, so if all None or a formula we can be roughly sure it's a null row
            if all(
                [
                    str(value) == "None"
                    or str(value).startswith("=")
                    or str(value) == " "
                    for value in record
                ]
            ):
                null_idx.append(i)

        for mv in tqdm(tab_mvs, f"removing null rows for {tab} tab"):
            # reversed to avoid deleting index 1 and then the index 2 becoming index 1 etc
            for i in sorted(null_idx, reverse=True):
                logger.debug(
                    f"deleting value | idx: {i} | tab: {mv.cell.tab} | name: {mv.cell.name} | value: {mv.values[i]} "
                )
                del mv.values[i]

        return True
    except Exception as e:
        logger.error(f"_remove_null_rows() unsuccessful for {tab}: {e}")
        return False


def _get_md_values_by_tab(
    metadata_values: Dict[str, MetadataValues], tab: str
) -> List[MetadataValues]:
    """gets all the MetadataValues objects from a tab

    Args:
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        tab (str): the tab to get the values from, pulled from the tab attribute of the
            MetadataCell within the MetadataValues object

    Returns:
        List[MetadataValues]: the subset of MetadataValues from that tab
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    return [v for v in metadata_values.values() if v.cell.tab == tab]


def _validate_metadata_values(
    metadata_values: Dict[str, MetadataValues], role_config: List[Tuple[str]]
) -> bool:
    """calls the validate method for each MetadataValues object

    Args:
        metadata_values (List[MetadataValues]): the list of extracted metadata_values

    Returns:
        bool: when all are validated
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    for mv in tqdm(metadata_values.values(), "validating cells"):

        if (mv.cell.tab, mv.cell.name) in role_config:
            mv.validate()
        else:
            # have to have empty dicts but they're initialised in validate
            # so skip assigns empty dicts to the hard and soft fails
            mv.skip()
    return True


def _get_all_fails_dicts(
    metadata_values: Dict[str, MetadataValues],
) -> Dict:
    """extracts the hard and soft fails from the metadata values

    Args:
        metadata_values (List[MetadataValues]): the list of extracted metadata_values

    Returns:
        Dict[str, Union[List, pd.DataFrame]]: dict of the fail_dfs
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    all_hard_fails, all_soft_fails = [], []

    for mv in tqdm(metadata_values.values(), "finding errors"):
        for k, v in mv.hard_fails.items():
            all_hard_fails.append(
                {
                    "tab": mv.cell.tab,
                    "name": mv.cell.name,
                    "value": k,
                    "reason": v,
                }
            )

        for k, v in mv.soft_fails.items():
            all_soft_fails.append(
                {
                    "tab": mv.cell.tab,
                    "name": mv.cell.name,
                    "value": k,
                    "reason": v,
                }
            )

    return {
        "hard_fails": all_hard_fails,
        "soft_fails": all_soft_fails,
    }


def _get_comparative_fails(
    metadata_values: Dict[str, MetadataValues], template_map: Dict
) -> List[Dict]:
    """wrapper to run all comparative validation checks

    Args:
        metadata_values (List[MetadataValues]): populated MetadataValues objects

    Returns:
        List[Dict]: list of errors in the same format as hard and soft fails
    """
    tabs = [
        "Dataset Resource",
        "Dataset File",
        "Dataset Series",
        "Variables",
        "Codes and Values",
    ]

    df_dict = {tab: md_values_to_df(metadata_values, tab) for tab in tabs}
    # this is the only place that uses pandas, it should be safe from pd version differences
    # these require dataframes as their interface which why they can be run
    # in a loop
    df_comparative_validations = {
        "Dataset File": [
            comp_val.check_DatasetFile_csv_string_identifier,
            comp_val.check_DatasetFile_multiple_files_to_append,
            comp_val.check_DatasetFile_csv_column_separators,
            comp_val.check_DatasetFile_csv_number_of_header_rows,
        ],
        "Variables": [
            comp_val.check_Variables_duplicate_variable_names,
            comp_val.check_Variables_variable_length_precision,
            comp_val.check_Variables_null_values_denoted_by,
            comp_val.check_Variables_foreign_key_file_name,
        ],
        "Codes and Values": [
            comp_val.check_CodesAndValues_duplicate_variable_names,
        ],
    }

    mv_comp_vals = [
        comp_val.check_table_names_appear_in_main_tabs,
        comp_val.check_all_unique_values_cols,
        comp_val.check_Variables_CodesAndValues_is_this_a_code,
        comp_val.check_DatasetSeries_length_for_DatasetResource_vs_bigquery_table,
    ]

    comp_errors = []

    for comp_func in mv_comp_vals:
        comp_errors.extend(comp_func(metadata_values, template_map))

    for tab_name, val_funcs in df_comparative_validations.items():
        for val_func in val_funcs:
            comp_errors.extend(val_func(df_dict[tab_name], template_map))

    return comp_errors


def _add_cell_references_to_df(
    wb: Workbook,
    metadata_values: Dict[str, MetadataValues],
    fails_list: List[Dict],
) -> pd.DataFrame:
    """adds the cell references to the cells with fails from the fails_list

    Args:
        wb (Workbook): the metadata template workbook (not modified)
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        fails_list (List[Dict]): list of individual fail dicts

    Returns:
        pd.DataFrame: dataframe of the fails with added cell references
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    fails_with_refs = []

    for fail in fails_list:
        fail["cell_refs"] = _get_fail_references(wb, metadata_values, fail)
        fails_with_refs.append(fail)

    return (
        pd.DataFrame(fails_with_refs).drop_duplicates().reset_index(drop=True)
    )


def _get_fail_references(
    wb: Workbook, metadata_values: Dict[str, MetadataValues], fail: Dict
) -> str:
    """_summary_

    Args:
        wb (Workbook): the metadata template workbook (not modified)
        metadata_values (List[MetadataValues]): the list of extracted metadata_values
        fail (Dict): single fail instance

    Returns:
        str: a joined str of the list of cell references
    """
    for key, val in locals().items():
        logger.debug(f"{key} = {compress_logging_value(val)}")
    mv = search_md_values(metadata_values, fail["tab"], fail["name"])

    # is len(mv.values) cleaner? avoids going over loads of empty rows
    # more risky but less cell refs for "None"s that go to the bottom of the sheet
    n_rows = len(mv.values)  # wb[mv.cell.tab].max_row - mv.cell.row_start
    values = []

    if mv.cell.column:
        for n in range(n_rows):
            cell_loc = f"{mv.cell.value_col}{mv.cell.row_start+n}"
            cell_value = wb[mv.cell.tab][cell_loc].value

            # convert the cell_value to str for easy comparison
            if str(cell_value) == str(fail["value"]):
                values.append(cell_loc)

    else:
        values.append(f"{mv.cell.value_col}{mv.cell.row_start}")

    values = ", ".join(values)
    logger.info(f"fail: {fail} | cell_locs: {values}")
    return values
