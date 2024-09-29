# import
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import os
import re
import warnings
import unicodedata
import pandas as pd
import csv
from openpyxl import load_workbook
from cryptography.fernet import Fernet
import json
import sys

class datareadr:
    def __init__(self):
        # main_window
        self.root = tk.Tk()
        self.root.title("DataReadr")
        # self.root.geometry("1000x600")
        self.screen_width = self.root.winfo_screenwidth()
        self.screen_height = self.root.winfo_screenheight()

        self.root.geometry(f'{self.screen_width}x{self.screen_height}')
        self.__import_file_data_options_page__()
        self.value_inside = tk.StringVar()
        self.value_inside_sep = tk.StringVar()
        self.working_directory = os.getcwd()
        self.file_name_path = sys.argv[0]

    def __load_ui__(self, path):
        self.file_path = path
        # ui_status_of_pages
        self.ui_status = "ok"
        self.__selecting_file_page_status__ = "ok"
        self.__preview_file_page_status__ = "ok"
        self.__insert_dataframe_page_status__ = "ok"
        self.__data_table_is_on = False
        self.code_json_writable = True
        self.code_json_importable = True

        self.root.mainloop()

    def __get_file_name__(self, file_path):
        self.file_name = f"{file_path}"
        os.path.split(self.file_name)
        self.output_name = os.path.split(self.file_name)[1]
        self.split_text = os.path.splitext(f"{self.output_name}")
        self.data_frame_name = self.split_text[0]
        return f"{self.data_frame_name}"

    def __normalize_text__(self, value, allow_unicode=False):
        value = str(value)
        if allow_unicode:
            value = unicodedata.normalize('NFKC', value)
        else:
            value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
        value = re.sub(r'[^\w\s-]', '', value.lower())
        modified_string = ''.join([i for i in value if not i.isdigit()])
        return re.sub(r'[-\s]+', '_', modified_string).strip('-_')

    def __open_file__(self, is_excel : bool, is_csv : bool, is_json : bool = False):
        if is_excel == True and is_csv == False and is_json == False:
            self.select_file_dialog = filedialog.askopenfilename(initialdir="/", title="Select An Excel File",
                                                                 filetypes=(("Excel Files", "*.xlsx"), ("xls", "*.xls")))
        elif is_excel == False and is_csv == True and is_json == False:
            self.select_file_dialog = filedialog.askopenfilename(initialdir="/", title="Select A CSV File",
                                                                 filetypes=[("CSV Files", "*.csv")])
        elif is_excel == False and is_csv == False and is_json == True:
            self.select_file_dialog = filedialog.askopenfilename(initialdir="/", title="Select A CSV File",
                                                                 filetypes=[("JSON Files", "*.json")])

        self.selected_file_path = f"{self.select_file_dialog}"
        self.selecting_file_entry_variable.set(f"{self.selected_file_path}")
        self.check_btn["text"] = "Check..."
        self.check_btn["command"] = lambda: self.__check_func__()
        self.code_preview_text.config(state=tk.NORMAL)
        self.code_preview_text.delete('1.0', tk.END)
        self.code_preview_text.config(state=tk.DISABLED)
        try:
            self.data_text.delete('1.0', tk.END)
        except:
            try:
                self.data_text.delete('1.0', tk.END)
            except:
                pass
        try:
            self.my_tree.destroy()
        except:
            pass
        self.__data_table_is_on = False

    def open_data(self):
        # grab the file
        try:
            # Create a dataframe
            if self.split_text[1] == ".xlsx" or self.split_text[1] == ".xls":
                self.df = pd.read_excel(self.selected_file_path,)
                # Clear the treeview
                self.my_tree.delete(*self.my_tree.get_children())

                # Get the Headers
                self.my_tree['column'] = list(self.df.columns)
                self.my_tree['show'] = 'headings'

                # Show the headers
                for col in self.my_tree['column']:
                    self.my_tree.heading(col, text=col, )

                # Show Data
                df_rows = self.df.to_numpy().tolist()
                for row in df_rows:
                    self.my_tree.insert("", "end", values=row)

                self.my_tree.grid(row=4, column=1, sticky="w", padx=35)
                xs = ttk.Scrollbar(self.root, orient=tk.VERTICAL, command=self.my_tree.yview)
                self.my_tree["xscrollcommand"] = xs.set
                xs.grid(row=4, column=2, sticky=(tk.S, tk.N,))

            elif self.split_text[1] == ".csv":
                with open(self.selected_file_path, 'r', newline='') as file:
                    csv_reader = csv.reader(file)
                    header = next(csv_reader)  # Read the header row
                    self.my_tree.delete(*self.my_tree.get_children())  # Clear the current data

                    self.my_tree["columns"] = header
                    for col in header:
                        self.my_tree.heading(col, text=col)
                        self.my_tree.column(col, width=100)

                    for row in csv_reader:
                        self.my_tree.insert("", "end", values=row)

                    self.my_tree.grid(row=4, column=1, sticky="w", padx=35)
                    xs = ttk.Scrollbar(self.root, orient=tk.VERTICAL, command=self.my_tree.yview)
                    self.my_tree["xscrollcommand"] = xs.set
                    xs.grid(row=4, column=2, sticky=(tk.S, tk.N,))

            elif self.split_text[1] == ".json":
                try:
                    with open(f"{self.selected_file_path}", "r") as data:
                        data = json.load(data)

                        formatted_json = json.dumps(data, indent=4)

                        self.data_text = tk.Text(self.root,  height=11, width=62, font=("Poppins Regular", 12))
                        self.data_text.grid(row=4, column=1, padx=45, sticky="w")
                        self.data_text.insert(tk.END, formatted_json)
                except:
                    self.code_json_importable = False
                    self.code_preview_text["state"] = tk.NORMAL
                    self.code_preview_text.delete('1.0', tk.END)
                    messagebox.showerror("Error", "There is an error with the file, try using another file.")

        except Exception as e:
            messagebox.showerror("Woah!", f'There was a problem! {e}')

    def __check_func__(self):
        self.code_preview_text.config(state=tk.NORMAL)
        self.data_frame_name_entry_variable.set(f"{self.__get_file_name__(f"{self.selected_file_path}")}")
        if self.split_text[1] == ".xlsx" or self.split_text[1] == ".xls":
            self.code_preview_code = \
f"""import datareadr as dr
{self.__normalize_text__(self.data_frame_name)} = dr.datareadr.recalling("{self.working_directory}\\{self.__get_file_name__(f"{self.selected_file_path}")}.dr").return_val()
print({self.__normalize_text__(self.data_frame_name)})
"""
        elif self.split_text[1] == ".csv":
            self.code_preview_code = \
f"""import datareadr as dr
{self.__normalize_text__(self.data_frame_name)} = dr.datareadr.recalling("{self.working_directory}\\{self.__get_file_name__(f"{self.selected_file_path}")}.dr").return_val()
print({self.__normalize_text__(self.data_frame_name)})
"""
        elif self.split_text[1] == ".json":
            if self.code_json_writable == True:
                self.code_preview_code = \
f"""import datareadr as dr
{self.__normalize_text__(self.data_frame_name)} = dr.datareadr.recalling("{self.working_directory}\\{self.__get_file_name__(f"{self.selected_file_path}")}.dr").return_val()
print({self.__normalize_text__(self.data_frame_name)})
                """
            else:
                pass
        self.code_preview_text.insert(tk.END, self.code_preview_code)
        self.check_btn["text"] = "Import..."
        self.check_btn["command"] = self.__import_func__
        self.code_preview_text.config(state=tk.DISABLED)
        self.__data_table_is_on = True
        if self.__data_table_is_on == True:
            self.__creating_data_preview_table__()
        else:
            pass

        try:
            self.wb = load_workbook(self.selected_file_path, read_only=True, keep_links=False)
            self.value_inside.set("Select Sheet")
            self.sheet_names_option = ttk.OptionMenu(self.import_options_frame, self.value_inside, *self.wb.sheetnames,)
            self.sheet_names_option.grid(row=1, column=1, pady=20)
        except:
            pass
        try:
            if self.split_text[1] == '.csv':
                self.value_inside_sep.set("Select Separator")
                values = ["Select Separator", "Comma", "Tab", "Space", "Semicolon"]
                self.separator_option = ttk.OptionMenu(self.import_options_frame, self.value_inside_sep,
                                                         *values,)
                self.separator_option.grid(row=2, column=1, padx=20)
            else:
                pass
        except:
            pass


    def reload_data(self):
        self.my_tree.delete(*self.my_tree.get_children())
        self.df = pd.read_excel(self.selected_file_path, sheet_name=f"{self.value_inside.get()}")
        # Clear the treeview
        self.my_tree.delete(*self.my_tree.get_children())

        # Get the Headers
        self.my_tree['column'] = list(self.df.columns)
        self.my_tree['show'] = 'headings'

        # Show the headers
        for col in self.my_tree['column']:
            self.my_tree.heading(col, text=col, )

        # Show Data
        df_rows = self.df.to_numpy().tolist()
        for row in df_rows:
            self.my_tree.insert("", "end", values=row)

    def __creating_data_preview_table__(self):
        # tree_view
        self.my_tree = ttk.Treeview(self.root, height=19, selectmode="extended")

        # hack the column height
        self.my_tree.heading('#0', text="\n")

        # ys = ttk.Scrollbar(self.root, orient=tk.HORIZONTAL, command=self.my_tree.yview)
        # self.my_tree["yscrollcommand"] = xs.set
        # ys.grid(row=1, column=0, sticky=(tk.E, tk.W))

        self.open_data()

    def __import_from_excel_page__(self):
        self.root.geometry(f'{self.screen_width-50}x{self.screen_height-85}')
        # self.root.wm_minsize(width=1650, height=880)
        # self.root.wm_maxsize(width=1650, height=880)
        self.root.title("DataReadr | Excel Import")
        self.root.grid_columnconfigure(0, weight=0)
        for i in self.root.winfo_children():
            i.destroy()

        # text_variables
        self.selecting_file_entry_variable = tk.StringVar()
        self.data_frame_name_entry_variable = tk.StringVar()

        # frames
        self.import_options_frame = tk.Frame(self.root, width=400, height=170, borderwidth=1, relief=tk.SOLID, )
        self.import_options_frame.grid(row=6, column=1, padx=20, sticky="w", )
        self.code_preview_frame = ttk.Frame(self.root, width=800, height=190, borderwidth=1, relief=tk.RIDGE)
        self.code_preview_frame.grid(row=6, column=1, padx=20, sticky=("ee",), )
        self.data_preview_frame = ttk.Frame(self.root, width=1650, height=420, borderwidth=1, relief=tk.RIDGE)
        self.root.rowconfigure(4, weight=2)
        self.root.columnconfigure(1, weight=1)
        self.data_preview_frame.grid(row=4, column=1, padx=0, sticky=("ew", "we"))

        # labels
        self.select_file_lbl = ttk.Label(self.root, text="File/URL:", font=("Poppins Regular", 10))
        self.code_preview_lbl = ttk.Label(self.root, text="Code Preview:", font=("Poppins Regular", 10))
        self.data_preview_lbl = ttk.Label(self.root, text="Data Preview:", font=("Poppins Regular", 10))
        self.import_options_lbl = ttk.Label(self.root, text="Import Options:", font=("Poppins Regular", 10))
        self.data_preview_lbl.grid(row=3, column=1, sticky="w", padx=20, pady=10)
        self.code_preview_lbl.grid(row=5, column=1, sticky="e", padx=20, pady=10)
        self.import_options_lbl.grid(row=5, column=1, sticky="w", padx=20, pady=10)
        self.select_file_lbl.grid(row=1, column=1, sticky="w", padx=20)
        self.space = ttk.Label(self.root, text="").grid(row=0)
        self.v_name_label = ttk.Label(master=self.import_options_frame, text="Name:",
                                      font=("Poppins Regular", 10)).grid(row=0, column=0, pady=10, padx=10)
        self.sheet_names_label = ttk.Label(master=self.import_options_frame, text="Sheet Name:",
                                       font=("Poppins Regular", 10)).grid(row=1, column=0, pady=10, padx=20)

        # buttons
        self.select_file_btn = ttk.Button(self.root, text="Browse...", width=10, command=lambda: self.__open_file__(is_excel=True, is_csv=False))
        self.check_btn = ttk.Button(self.root, text="Check...", width=15, command=lambda: self.__check_func__())
        self.back_btn = ttk.Button(self.root, text="Back...", width=15,
                                   command=lambda: self.__import_file_data_options_page__())
        self.reload_btn = ttk.Button(self.root, text="Reload...", width=10, command= lambda : self.reload_data())
        self.reload_btn.grid(row=3, column=1, sticky="e", pady=10, padx=20)
        self.back_btn.grid(row=7, column=1, padx=20, sticky="w")
        self.check_btn.grid(row=7, column=1, sticky="e", pady=20, padx=20)
        self.select_file_btn.grid(row=1, column=1, sticky="e", padx=20)

        # entry
        self.v_name_entry = ttk.Entry(self.import_options_frame, width=50,
                                      textvariable=self.data_frame_name_entry_variable).grid(row=0, column=1, padx=20)
        self.entry_file_selecting = ttk.Entry(self.root, width=120, font=("Poppins", 12),
                                              textvariable=self.selecting_file_entry_variable)
        self.entry_file_selecting.grid(row=2, column=1, padx=20, sticky="we")

        # text_area
        self.code_preview_text = tk.Text(self.code_preview_frame, height=4.49, width=62, font=("Poppins Regular", 12), )
        self.code_preview_text.config(state=tk.DISABLED)
        self.code_preview_text.grid(row=0, column=1, sticky="s")

    def __import_from_csv_page__(self):
        self.root.geometry(f'{self.screen_width-50}x{self.screen_height-85}')
        # self.root.wm_minsize(width=1650, height=880)
        # self.root.wm_maxsize(width=1650, height=880)
        self.root.title("DataReadr | CSV Import")
        self.root.grid_columnconfigure(0, weight=0)
        for i in self.root.winfo_children():
            i.destroy()

        # text_variables
        self.selecting_file_entry_variable = tk.StringVar()
        self.data_frame_name_entry_variable = tk.StringVar()

        # frames
        self.import_options_frame = tk.Frame(self.root, width=400, height=170, borderwidth=1, relief=tk.SOLID, )
        self.import_options_frame.grid(row=6, column=1, padx=20, sticky="w", )
        self.code_preview_frame = ttk.Frame(self.root, width=800, height=190, borderwidth=1, relief=tk.RIDGE)
        self.code_preview_frame.grid(row=6, column=1, padx=20, sticky="e", )
        self.data_preview_frame = ttk.Frame(self.root, width=1650, height=420, borderwidth=1, relief=tk.RIDGE)
        self.root.rowconfigure(4, weight=2)
        self.root.columnconfigure(1, weight=1)
        self.data_preview_frame.grid(row=4, column=1, padx=0, sticky=("ew", "we"))

        # labels
        self.select_file_lbl = ttk.Label(self.root, text="File/URL:", font=("Poppins Regular", 10))
        self.code_preview_lbl = ttk.Label(self.root, text="Code Preview:", font=("Poppins Regular", 10))
        self.data_preview_lbl = ttk.Label(self.root, text="Data Preview:", font=("Poppins Regular", 10))
        self.import_options_lbl = ttk.Label(self.root, text="Import Options:", font=("Poppins Regular", 10))
        self.data_preview_lbl.grid(row=3, column=1, sticky="w", padx=20, pady=10)
        self.code_preview_lbl.grid(row=5, column=1, sticky="e", padx=700, pady=10)
        self.import_options_lbl.grid(row=5, column=1, sticky="w", padx=20, pady=10)
        self.select_file_lbl.grid(row=1, column=1, sticky="w", padx=20)
        self.space = ttk.Label(self.root, text="").grid(row=0)
        self.v_name_label = ttk.Label(master=self.import_options_frame, text="Name:",
                                      font=("Poppins Regular", 10)).grid(row=0, column=0, pady=10, padx=10)
        self.v_name2_label = ttk.Label(master=self.import_options_frame, text="Seprator:",
                                       font=("Poppins Regular", 10)).grid(row=2, column=0, pady=10)

        # buttons
        self.select_file_btn = ttk.Button(self.root, text="Browse...", width=10,
                                          command=lambda: self.__open_file__(is_excel=False, is_csv=True))
        self.check_btn = ttk.Button(self.root, text="Check...", width=15, command=lambda: self.__check_func__())
        self.back_btn = ttk.Button(self.root, text="Back...", width=15, command=lambda: self.__import_file_data_options_page__())
        self.back_btn.grid(row=7, column=1, padx=20, sticky="w")
        self.check_btn.grid(row=7, column=1, sticky="e", pady=20, padx=20)
        self.select_file_btn.grid(row=1, column=1, sticky="e", padx=20)

        # entry
        self.v_name_entry = ttk.Entry(self.import_options_frame, width=50,
                                      textvariable=self.data_frame_name_entry_variable).grid(row=0, column=1, padx=20)
        self.entry_file_selecting = ttk.Entry(self.root, width=120, font=("Poppins", 12),
                                              textvariable=self.selecting_file_entry_variable)
        self.entry_file_selecting.grid(row=2, column=1, padx=20, sticky="we")

        # text_area
        self.code_preview_text = tk.Text(self.code_preview_frame, height=4.49, width=62, font=("Poppins Regular", 12))
        self.code_preview_text.config(state=tk.DISABLED)
        self.code_preview_text.grid(row=0, column=1, sticky="s")

    def __import_from_json_page__(self):
        self.root.geometry(f'{self.screen_width-50}x{self.screen_height-85}')
        # self.root.wm_minsize(width=1650, height=880)
        # self.root.wm_maxsize(width=1650, height=880)
        self.root.title("DataReadr | JSON Import")
        self.root.grid_columnconfigure(0, weight=0)
        for i in self.root.winfo_children():
            i.destroy()

        # text_variables
        self.selecting_file_entry_variable = tk.StringVar()
        self.data_frame_name_entry_variable = tk.StringVar()

        # frames
        self.import_options_frame = tk.Frame(self.root, width=400, height=170, borderwidth=1, relief=tk.SOLID, )
        self.import_options_frame.grid(row=6, column=1, padx=20, sticky="w", )
        self.code_preview_frame = ttk.Frame(self.root, width=800, height=190, borderwidth=1, relief=tk.RIDGE)
        self.code_preview_frame.grid(row=6, column=1, padx=20, sticky="e", )
        self.data_preview_frame = ttk.Frame(self.root, width=1650, height=420, borderwidth=1, relief=tk.RIDGE)
        self.root.rowconfigure(4, weight=2)
        self.root.columnconfigure(1, weight=1)
        self.data_preview_frame.grid(row=4, column=1, padx=0, sticky=("ew", "we"))

        # labels
        self.select_file_lbl = ttk.Label(self.root, text="File/URL:", font=("Poppins Regular", 10))
        self.code_preview_lbl = ttk.Label(self.root, text="Code Preview:", font=("Poppins Regular", 10))
        self.data_preview_lbl = ttk.Label(self.root, text="Data Preview:", font=("Poppins Regular", 10))
        self.import_options_lbl = ttk.Label(self.root, text="Import Options:", font=("Poppins Regular", 10))
        self.data_preview_lbl.grid(row=3, column=1, sticky="w", padx=20, pady=10)
        self.code_preview_lbl.grid(row=5, column=1, sticky="e", padx=700, pady=10)
        self.import_options_lbl.grid(row=5, column=1, sticky="w", padx=20, pady=10)
        self.select_file_lbl.grid(row=1, column=1, sticky="w", padx=20)
        self.space = ttk.Label(self.root, text="").grid(row=0)
        self.v_name_label = ttk.Label(master=self.import_options_frame, text="Name:",
                                      font=("Poppins Regular", 10)).grid(row=0, column=0, pady=10, padx=10)
        # buttons
        self.select_file_btn = ttk.Button(self.root, text="Browse...", width=10,
                                          command=lambda: self.__open_file__(is_excel=False, is_csv=False, is_json=True))
        self.check_btn = ttk.Button(self.root, text="Check...", width=15, command=lambda: self.__check_func__())
        self.back_btn = ttk.Button(self.root, text="Back...", width=15, command=lambda: self.__import_file_data_options_page__())
        self.back_btn.grid(row=7, column=1, padx=20, sticky="w")
        self.check_btn.grid(row=7, column=1, sticky="e", pady=20, padx=20)
        self.select_file_btn.grid(row=1, column=1, sticky="e", padx=20)

        # entry
        self.v_name_entry = ttk.Entry(self.import_options_frame, width=50,
                                      textvariable=self.data_frame_name_entry_variable).grid(row=0, column=1, padx=20)

        self.entry_file_selecting = ttk.Entry(self.root, width=120, font=("Poppins", 12),
                                              textvariable=self.selecting_file_entry_variable)
        self.entry_file_selecting.grid(row=2, column=1, padx=20, sticky="we")

        # text_area
        self.code_preview_text = tk.Text(self.code_preview_frame, height=4.49, width=62, font=("Poppins Regular", 12))
        self.code_preview_text.config(state=tk.DISABLED)
        self.code_preview_text.grid(row=0, column=1, sticky="s")

    def __import_file_data_options_page__(self):
        self.root.grid_columnconfigure(0, weight=2)
        self.root.title("DataReadr")
        # self.root.geometry("1000x600")
        self.root.geometry(f'{self.screen_width-750}x{self.screen_height-250}')
        for i in self.root.winfo_children():
            i.destroy()
        # self.root.grid_columnconfigure(1, weight=1)
        # self.root.grid_columnconfigure(2, weight=1)
        # self.root.grid_columnconfigure(3, weight=1)

        self._welcome_lbl_ = ttk.Label(self.root, text="What Do You Want To Import?", font=("Poppins Regular", 24, "bold")).grid(row=0,pady=40, padx=20)
        s = ttk.Style()
        s.configure('my.TButton', font=('Poppins Regular', 20))
        self.import_excel_file_btn = ttk.Button(self.root, text="From Excel", style="my.TButton", width=20,command = lambda : self.__import_from_excel_page__())
        self.import_excel_file_btn.grid(row=1, pady=10)
        self.import_csv_file_btn = ttk.Button(self.root, text="From CSV", style="my.TButton", width=20, command = lambda : self.__import_from_csv_page__())
        self.import_csv_file_btn.grid(row=2, pady=10)
        self.import_json_file_btn = ttk.Button(self.root, text="From JSON", style="my.TButton", width=20, command = lambda : self.__import_from_json_page__())
        self.import_json_file_btn.grid(row=3, pady=10)
    class recalling(pd.DataFrame):
        def __init__(self, path, *args, **kwargs, ):
            super().__init__(*args, **kwargs)
            split_tup = os.path.splitext(f'{self.read(f'{path}')}')
            if split_tup[1] == ".xlsx" or split_tup[1] == ".xls":
                self.df = pd.read_excel(f"{self.read(f'{path}')}")
            elif split_tup[1] == ".csv":
                self.df = pd.read_csv(f"{self.read(f'{path}')}")
            elif split_tup[1] == ".json":
                self.df = pd.read_json(f"{self.read(f'{path}')}")
            self.my_df = self.df.copy()
            self.my_df = self.my_df.drop(0)

        def read(self, file):
            with open(f"{file}", "r+") as f:
                self.data = f.read()
            return self.data

        def return_val(self):
            warnings.filterwarnings("ignore",
                                    "Pandas doesn't allow columns to be created via a new attribute name - see https://pandas.pydata.org/pandas-docs/stable/indexing.html#attribute-access",
                                    UserWarning)
            return (self.my_df)

    def __import_func__(self):
            try:
                if self.split_text[1] == '.json':
                    if self.code_json_importable == True:
                        with open(f"{self.file_path}", "a") as self.__current_operating_file__:
                            self.__current_operating_file__.write(f"\n{self.code_preview_code}\n")
                        messagebox.showinfo("Info", "Successfully Imported Dataframe")
                        exit()
                    else:
                        messagebox.showerror("Error", "Invalid File.")
                        self.code_preview_text.delete('1.0', tk.END)
                elif self.split_text[1] == ".csv":
                    self.sep = ""
                    if self.value_inside_sep.get() == "Comma":
                        self.sep = ","
                    elif self.value_inside_sep.get() == "Semicolon":
                        self.sep = ";"
                    elif self.value_inside_sep.get() == "Tab":
                        self.sep = "t"
                    elif self.value_inside_sep.get() == "Space":
                        self.sep = "  "
                    self.code_preview_code = f"""
import pandas as pd
{self.__normalize_text__(self.data_frame_name)} = pd.read_csv('{self.selected_file_path}', sep = '{self.sep}')
print({self.__normalize_text__(self.data_frame_name)})

"""
                    with open(f"{self.file_path}", "a") as self.__current_operating_file__:
                        self.__current_operating_file__.write(f"\n{self.code_preview_code}\n")

                    messagebox.showinfo("Info", "Successfully Imported Dataframe")
                    exit()
                else:
                    with open(f"{self.file_path}", "a") as self.__current_operating_file__:
                        self.__current_operating_file__.write(f"\n{self.code_preview_code}\n")
                    with open(f"{self.working_directory}/{self.__get_file_name__(f"{self.selected_file_path}")}.dr", "w+") as df:
                        df.write(f'{self.selected_file_path}')

                    messagebox.showinfo("Info", "Successfully Imported Dataframe")
                    exit()
            except Exception as e:
                print(f"There was an error!", e)
                messagebox.showerror("Error", "We encountered en error ", e)

