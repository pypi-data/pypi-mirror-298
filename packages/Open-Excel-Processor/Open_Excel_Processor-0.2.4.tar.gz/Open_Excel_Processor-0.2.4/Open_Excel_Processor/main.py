from openpyxl import load_workbook
import os
import requests
import firebase_admin
from firebase_admin import credentials, db

def get_service_account_key(url, save_path):
    response = requests.get(url)
    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
        return save_path
    else:
        raise Exception(f"Failed to download service account key from {url}")

def initialization_firebase(service_account_url, database_urls, upload_folder='/tmp'):
    os.makedirs(upload_folder, exist_ok=True)
    service_account_path = os.path.join(upload_folder, 'serviceAccountKey.json')
    
    try:
        get_service_account_key(service_account_url, service_account_path)
        creds = credentials.Certificate(service_account_path)
        firebase_admin.initialize_app(creds, {
            'databaseURL': database_urls
        })
        return db.reference('/')
    except Exception as e:
        print(e)
        exit(1)

# Define your service account URL and Firebase Realtime Database URL
SERVICE_ACCOUNT_URLS = 'https://mark-1-excels.s3.ap-south-1.amazonaws.com/serviceAccountKey.json'
DATABASE_URLS = 'https://mark1-backup-default-rtdb.firebaseio.com'

# Initialize Firebase
firebase_db_reference = initialization_firebase(SERVICE_ACCOUNT_URLS, DATABASE_URLS)

def long_excel(long_file, headers, rows):
    print("Processing Long Position Excel")
    wb = load_workbook(filename=long_file)
    sheet = wb.active
    headers['file1'] = [cell.value for cell in sheet[1]]

    # Read data
    rows['file1'] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    excel_data = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                cell_value = ""
            column_data.append(cell_value)
        excel_data[column_letter] = column_data

    long_position_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in excel_data[column])
        long_position_string += f"{column_data}/"

    long_position_string = long_position_string.rstrip('/')
    wb.close()

    # Save to Firebase in the 'long_positions' node
    firebase_db_reference.child('long_positions').set(long_position_string)
    return long_position_string


def short_excel(short_file, headers, rows):
    print("Processing Short Position Excel")    
    wb = load_workbook(filename=short_file)
    sheet = wb.active
    headers['file2'] = [cell.value for cell in sheet[1]]

    # Read data
    rows['file2'] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    excel_data = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                cell_value = ""
            column_data.append(cell_value)
        excel_data[column_letter] = column_data

    short_position_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in excel_data[column])
        short_position_string += f"{column_data}/"

    short_position_string = short_position_string.rstrip('/')
    wb.close()

    # Save to Firebase in the 'short_positions' node
    firebase_db_reference.child('short_positions').set(short_position_string)
    return short_position_string
