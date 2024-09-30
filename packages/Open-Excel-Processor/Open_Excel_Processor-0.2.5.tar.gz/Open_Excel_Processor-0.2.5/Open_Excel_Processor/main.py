from openpyxl import load_workbook
import os
import requests
import firebase_admin
from firebase_admin import credentials, db

def download_service_account_key(url, save_path):
    """Download the service account key from the given URL."""
    response = requests.get(url)
    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
        return save_path
    else:
        raise Exception(f"Failed to download service account key from {url}")

def initialize_firebase(service_account_url, database_url, upload_folder='/tmp', app_name='default'):
    """Initialize Firebase app with the service account key and database URL."""
    os.makedirs(upload_folder, exist_ok=True)
    service_account_path = os.path.join(upload_folder, 'serviceAccountKey.json')
    
    try:
        download_service_account_key(service_account_url, service_account_path)
        creds = credentials.Certificate(service_account_path)
        
        # Check if the app is already initialized
        if not firebase_admin._apps:
            firebase_admin.initialize_app(creds, {'databaseURL': database_url}, name=app_name)
        else:
            print(f"Firebase app '{app_name}' is already initialized.")
        
        return db.reference('/')
    except Exception as e:
        print(f"Error initializing Firebase: {e}")
        exit(1)

# Service account URL and Firebase Realtime Database URL
SERVICE_ACCOUNT_URL = 'https://mark-1-excels.s3.ap-south-1.amazonaws.com/serviceAccountKey.json'
DATABASE_URL = 'https://mark1-backup-default-rtdb.firebaseio.com'

# Initialize Firebase and get database reference
firebase_db_reference = initialize_firebase(SERVICE_ACCOUNT_URL, DATABASE_URL)

def process_excel(file_path, headers, rows, excel_type):
    """Generic function to process Excel files."""
    print(f"Processing {excel_type} Position Excel")
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    headers[excel_type] = [cell.value for cell in sheet[1]]

    # Read data
    rows[excel_type] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    excel_data = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            column_data.append(cell_value if cell_value else "")
        excel_data[column_letter] = column_data

    # Create a string to store the final result
    final_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in excel_data[column])
        final_string += f"{column_data}/"

    final_string = final_string.rstrip('/')
    wb.close()

    # Save to Firebase in the appropriate node
    firebase_db_reference.child(f'{excel_type}_positions').set(final_string)
    return final_string


def process_long_excel(long_file, headers, rows):
    """Process the long position Excel file."""
    return process_excel(long_file, headers, rows, 'long')


def process_short_excel(short_file, headers, rows):
    """Process the short position Excel file."""
    return process_excel(short_file, headers, rows, 'short')
