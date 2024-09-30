from openpyxl import load_workbook
import os
import requests
import firebase_admin
from firebase_admin import credentials, db

def download_service_account_key(url, save_path):
    response = requests.get(url)
    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
        return save_path
    else:
        raise Exception(f"Failed to download service account key from {url}")

def initialize_firebase(service_account_url, database_url, upload_folder='/tmp'):
    os.makedirs(upload_folder, exist_ok=True)
    service_account_path = os.path.join(upload_folder, 'serviceAccountKey.json')
    
    try:
        download_service_account_key(service_account_url, service_account_path)
        cred = credentials.Certificate(service_account_path)
        firebase_admin.initialize_app(cred, {
            'databaseURL': database_url
        })
        return db.reference('/')
    except Exception as e:
        print(e)
        exit(1)

# Define your service account URL and Firebase Realtime Database URL
SERVICE_ACCOUNT_URL = 'https://mark-1-excels.s3.ap-south-1.amazonaws.com/serviceAccountKey.json'
DATABASE_URL = 'https://mark1-backup-default-rtdb.firebaseio.com'

# Initialize Firebase
firebase_ref = initialize_firebase(SERVICE_ACCOUNT_URL, DATABASE_URL)

# Firebase configuration details for REST API
apiKey = "AIzaSyBYlRzXObxn7d5oVWzgGqXhM6xaBFkXN58",
authDomain = "mark1-backup.firebaseapp.com",
databaseURL = "https://mark1-backup-default-rtdb.firebaseio.com",
projectId = "mark1-backup",
storageBucket = "mark1-backup.appspot.com",
messagingSenderId = "1070285792269",
appId = "1:1070285792269:web:8fb26aeb3cf0c158f2a144",
measurementId = "G-BSH0HVMJK8"

ref_all_data = firebase_admin.firebase_ref.child('all_data')

def long_excel(long_file,headers,rows):
    print("Long Position Excel")
    wb = load_workbook(filename=long_file)
    sheet = wb.active
    headers['file1'] = [cell.value for cell in sheet[1]]

    # Read data
    rows['file1'] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    exceldata = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                cell_value = ""
            column_data.append(cell_value)
        exceldata[column_letter] = column_data

    long_final_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in exceldata[column])
        final_string += f"{column_data}/"

    final_string = final_string.rstrip('/')
    wb.close()
    ref_all_data.set(long_final_string)
    return long_final_string


def short_excel(short_file,headers,rows):
    print("Long Position Excel")    
    wb = load_workbook(filename=short_file)
    sheet = wb.active
    headers['file2'] = [cell.value for cell in sheet[1]]

    # Read data
    rows['file2'] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    exceldata = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                cell_value = ""
            column_data.append(cell_value)
        exceldata[column_letter] = column_data

    short_final_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in exceldata[column])
        final_string += f"{column_data}/"

    short_final_string = final_string.rstrip('/')
    wb.close()
    ref_all_data.set(short_final_string)
    return short_final_string