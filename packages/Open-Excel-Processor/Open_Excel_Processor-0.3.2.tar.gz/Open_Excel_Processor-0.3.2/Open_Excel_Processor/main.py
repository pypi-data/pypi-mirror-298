from openpyxl import load_workbook


def process_long_excel(long_file, headers, rows):
    """Process the long position Excel file."""
    wb = load_workbook(filename=long_file)
    sheet = wb.active
    headers['file1'] = [cell.value for cell in sheet[1]]

    # Read data
    rows['file1'] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    exceldata = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                cell_value = ""
            column_data.append(cell_value)
        exceldata[column_letter] = column_data

    long_final_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in exceldata[column])
        long_final_string += f"{column_data}/"

    long_final_string = long_final_string.rstrip('/')
    wb.close()
    return long_final_string

def process_short_excel(short_file, headers, rows):
    """Process the short position Excel file."""
    wb = load_workbook(filename=short_file)
    sheet = wb.active
    headers['file1'] = [cell.value for cell in sheet[1]]

    # Read data
    rows['file1'] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    exceldata = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                cell_value = ""
            column_data.append(cell_value)
        exceldata[column_letter] = column_data

    short_final_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in exceldata[column])
        final_string += f"{column_data}/"

    short_final_string = short_final_string.rstrip('/')
    wb.close()

    return short_final_string
