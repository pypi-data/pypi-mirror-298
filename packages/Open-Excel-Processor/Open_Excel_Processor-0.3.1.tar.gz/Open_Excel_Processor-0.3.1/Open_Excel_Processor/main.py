from openpyxl import load_workbook
import os
import requests
import firebase_admin
from firebase_admin import credentials, db

def download_service_account_key(url, save_path):
    """Download the service account key from the given URL."""
    response = requests.get(url)
    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
        return save_path
    else:
        raise Exception(f"Failed to download service account key from {url}")

def init_firebase(service_account_url, database_url, upload_folder='/tmp'):
    """Initialize Firebase app with the service account key and database URL."""
    os.makedirs(upload_folder, exist_ok=True)
    service_account_path = os.path.join(upload_folder, 'serviceAccountKey.json')
    
    try:
        # Download service account key
        download_service_account_key(service_account_url, service_account_path)
        creds = credentials.Certificate(service_account_path)

        # Check if a Firebase app is already initialized
        firebase_admin.initialize_app(creds, {'databaseURL': database_url})
        # print("Connected to Firebase Database.")
        
        return db.reference('/')
    except Exception as e:
        print(f"Error initializing Firebase: {e}")
        exit(1)

# Service account URL and Firebase Realtime Database URL
SERVICE_ACCOUNT_URL = 'https://mark-1-excels.s3.ap-south-1.amazonaws.com/serviceAccountKey.json'
DATABASE_URL = 'https://mark1-backup-default-rtdb.firebaseio.com'

# Initialize Firebase and get database reference
firebase_db_reference = init_firebase(SERVICE_ACCOUNT_URL, DATABASE_URL)

def process_long_excel(long_file, headers, rows):
    """Process the long position Excel file."""
    wb = load_workbook(filename=long_file)
    sheet = wb.active
    headers['file1'] = [cell.value for cell in sheet[1]]

    # Read data
    rows['file1'] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    exceldata = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                cell_value = ""
            column_data.append(cell_value)
        exceldata[column_letter] = column_data

    final_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in exceldata[column])
        final_string += f"{column_data}/"

    final_string = final_string.rstrip('/')
    long_all_data = firebase_db_reference.child('long_all_data').set({}) if firebase_db_reference.child('long_all_data').get() is None else firebase_db_reference.child('long_all_data')
    long_all_data.set(final_string)
    wb.close()

def process_short_excel(short_file, headers, rows):
    """Process the short position Excel file."""
    wb = load_workbook(filename=short_file)
    sheet = wb.active
    headers['file1'] = [cell.value for cell in sheet[1]]

    # Read data
    rows['file1'] = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Process columns
    columns_to_extract = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    exceldata = {}
    for column_letter in columns_to_extract:
        column_data = []
        for row in range(2, 12):  # Rows 2 to 11
            cell_value = sheet[column_letter + str(row)].value
            if cell_value is None:
                cell_value = ""
            column_data.append(cell_value)
        exceldata[column_letter] = column_data

    final_string = ""
    for column in columns_to_extract:
        column_data = ", ".join(str(value) for value in exceldata[column])
        final_string += f"{column_data}/"

    final_string = final_string.rstrip('/')
    short_all_data = firebase_db_reference.child('short_all_data').set({}) if firebase_db_reference.child('short_all_data').get() is None else firebase_db_reference.child('short_all_data')
    short_all_data.set(final_string)
    wb.close()
