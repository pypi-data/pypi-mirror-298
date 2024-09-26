# -*- coding: utf-8 -*-
import openpyxl
import os


class ExcelObj():
    def __init__(self, excel_name,table_name="Sheet1"):
        self.file_name = excel_name
        if not os.path.isfile(excel_name):
            self._create_exc()
        else:
            self._load_exc(excel_name, table_name)

    def init_column(self,column_list):
        self.work_table.append(column_list)
        self.work_book.save(self.file_name)

    def _create_exc(self):
        self.work_book = openpyxl.Workbook()
        self.work_table = self.work_book.active
        self.work_table.title = "Sheet1"

    def _load_exc(self, exc_name, table_name):
        self.work_book = openpyxl.load_workbook(exc_name)
        self.select_table_by_name(table_name)

    def select_table_by_name(self, table_name):
        table_list = self.work_book.sheetnames
        for table in table_list:
            if table_name == table:
                self.work_table = self.work_book[table_name]
                return
        raise Exception(f"{table_name} 没有找到此表")

    def _get_table_datas(self):
        data_list = []
        for row in self.work_table.values:
            data_list.append(row)
        return data_list

    def get_row_num(self):
        return self.work_table.max_row

    def get_column_num(self):
        return self.work_table.max_column

    def update_value_with_ABC(self, row, column, value):
        self._set_value(row, column, value)

    def update_value_with_position(self,row,column,value):
        ABCPosition = self.work_table.cell(row, column).coordinate
        self.work_table[ABCPosition] = value

    def _set_value(self, row, column, value):
        self.work_table[f"{column}{row}"] = value

    def _find(self,value):
        max_row = self.get_row_num()
        max_column = self.get_column_num()
        for i in range(1,max_row):
            for j in range(1,max_column):
                cel = self.work_table.cell(row=i, column=j)
                if cel.value == value:
                    return cel.coordinate

    def append_one(self, item):
        self.work_table.append(item)

    def save(self):
        self.work_book.save(self.file_name)

    def save_datas(self, data_list):
        for i in data_list:
            self.work_table.append(i)
        self.work_book.save(self.file_name)
        self.work_book.close()
        return True