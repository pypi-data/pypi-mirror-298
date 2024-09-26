import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

from .observer import Observer


class ExcelHandler(Observer):
    def __init__(self):
        super().__init__()
        self.df = None
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.workbook.remove(self.worksheet)  # 기본 시트 삭제
        self.sheet_names = set()

    def update(self, dataframe):
        self.df = dataframe
        self._update_worksheet()

    def _update_worksheet(self):
        self.worksheet.delete_rows(1, self.worksheet.max_row)
        for row in dataframe_to_rows(self.df, index=False, header=True):
            # self.worksheet.append(row)
            row = [self._convert_to_string_if_needed(cell) for cell in row]
            self.worksheet.append(row)

    def _convert_to_string_if_needed(self, value):
        """리스트나 배열을 문자열로 변환하여 엑셀에 넣을 수 있도록 처리"""
        if isinstance(value, (list, np.ndarray)):  # 리스트나 numpy 배열일 경우
            return ', '.join(map(str, value))  # 쉼표로 구분된 문자열로 변환
        return value

    def set_column_width(self, **kwargs):
        """컬럼 너비 설정"""
        for column, width in kwargs.items():
            if column in self.df.columns:
                col_idx = self.df.columns.get_loc(column) + 1
                col_letter = chr(64 + col_idx)
                self.worksheet.column_dimensions[col_letter].width = width
            else:
                pass

    def freeze_first_row(self):
        """첫 번째 행 고정"""
        self.worksheet.freeze_panes = self.worksheet["A2"]

    def enable_autowrap(self):
        """자동 줄 바꿈 활성화"""
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    def add_hyperlink(self, cell, url, display=None):
        """셀에 하이퍼링크 추가"""
        self.worksheet[cell].hyperlink = url
        self.worksheet[cell].value = display if display else url
        self.worksheet[cell].style = "Hyperlink"

    def add_hyperlinks_to_column(self, column_name, urls, display_texts=None):
        """컬럼에 있는 각 셀에 하이퍼링크 추가"""
        if len(urls) != len(self.df):
            raise ValueError(
                "The length of the URL list must match the length of the dataframe."
            )

        if display_texts and len(display_texts) != len(self.df):
            raise ValueError(
                "The length of the display_texts list must match the length of the dataframe."
            )

        col_idx = self.df.columns.get_loc(column_name) + 1
        for i, url in enumerate(urls, start=2):  # start=2 to account for header row
            cell = f"{chr(64 + col_idx)}{i}"
            display = display_texts[i - 2] if display_texts else None
            self.add_hyperlink(cell, url, display)

    def save(self, filename):
        """엑셀 파일 저장"""
        self.workbook.save(filename)

    def close(self):
        """엑셀 파일 닫기"""
        del self.workbook

    def __del__(self):
        self.close()

    # 시트 관련 기능 추가
    def create_sheet(self, dataframe, title=None):
        """새 시트 생성 및 데이터프레임 추가"""
        if title:
            if title in self.sheet_names:
                print(f"{title} sheet은 이미 존재합니다, 이름을 바꿔주세요")
                return False
            else:
                self.sheet_names.add(title)
        self.worksheet = self.workbook.create_sheet(title=title)
        self.df_to_worksheet(dataframe)
        self.df = dataframe
        return True

    def select_sheet(self, title):
        """시트 선택"""
        self.worksheet = self.workbook[title]
        data = list(self.worksheet.values)
        cols = data[0]  # First row as columns
        self.df = pd.DataFrame(data[1:], columns=cols)  # Use the rest as data
        return self.df

    def remove_sheet(self, title):
        """시트 제거"""
        sheet_to_remove = self.workbook[title]
        self.workbook.remove(sheet_to_remove)

    def df_to_worksheet(self, dataframe):
        """데이터프레임을 현재 워크시트에 추가"""
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            self.worksheet.append(r)

    def list_sheets(self):
        """시트 목록 반환"""
        return self.workbook.sheetnames

    def get_active_sheet(self):
        """현재 선택된 시트 이름 반환"""
        return self.worksheet.title
