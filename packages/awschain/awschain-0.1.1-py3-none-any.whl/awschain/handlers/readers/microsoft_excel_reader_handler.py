import os
import json
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from ..abstract_handler import AbstractHandler
from PIL import Image
import io

class MicrosoftExcelReaderHandler(AbstractHandler):
    
    def handle(self, request: dict) -> dict:
        print("Processing XLSX file...")

        # Initialize variables
        text_content = ''
        media_files = []

        # Load the workbook from the path specified in the request
        workbook_path = request.get("path", None)
        workbook = load_workbook(workbook_path, data_only=True)

        # Determine the write file path and output folder
        write_file_path = request.get("write_file_path", None)
        if write_file_path:
            base_folder = os.path.dirname(write_file_path)
            file_id = os.path.basename(write_file_path).split('.')[0]
            output_folder = os.path.join(base_folder, file_id)
            media_folder = os.path.join(output_folder, "media")
            os.makedirs(media_folder, exist_ok=True)
        
        # Initialize the metadata dictionary
        if "metadata" not in request:
            request["metadata"] = {}

        # Iterate through each sheet in the workbook
        for sheet in workbook:
            text_content += f"Sheet: {sheet.title}\n"
            
            # Extract text from cells
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text_content += str(cell) + '\t'
                text_content += '\n'
            
            if request.get("extract_media", False):
                # Extract images
                for image in sheet._images:
                    image_file_name = f"{sheet.title}_image_{image.anchor._from.col}{image.anchor._from.row}.png"
                    image_file_path = os.path.join(media_folder, image_file_name)
                    
                    # Handle image data correctly
                    image_stream = io.BytesIO(image._data())
                    pil_image = Image.open(image_stream)
                    pil_image.save(image_file_path)
                    
                    media_files.append({"type": "image", "path": image_file_path, "sheet": sheet.title})

        # Save the transcript to a text file
        transcript_file_path = os.path.join(output_folder, 'transcript.txt')
        with open(transcript_file_path, 'w') as transcript_file:
            transcript_file.write(text_content)

        # Create metadata with traceability for graph database
        metadata = {
            "original_file": workbook_path,
            "transcript_file": transcript_file_path,
            "media_files": media_files
        }
        
        # Save metadata to a json file
        metadata_file_path = os.path.join(output_folder, 'metadata.json')
        with open(metadata_file_path, 'w') as metadata_file:
            json.dump(metadata, metadata_file, indent=4)

        # Update the request metadata with the new metadata
        request["metadata"][workbook_path] = metadata

        # Update the request with the extracted text
        request.update({"text": text_content})

        # Call the next handler in the chain
        return super().handle(request)
