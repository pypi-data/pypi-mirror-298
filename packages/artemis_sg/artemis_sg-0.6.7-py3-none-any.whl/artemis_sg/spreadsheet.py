import logging
import math
import os
import re
from copy import copy
from inspect import getsourcefile

import isbnlib
from googleapiclient.discovery import build
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from PIL import Image as PIL_Image
from PIL import UnidentifiedImageError

from artemis_sg import app_creds, items, vendor
from artemis_sg.config import CFG

MODULE = os.path.splitext(os.path.basename(__file__))[0]


def get_worksheet(wb_obj, worksheet):
    ws = wb_obj.worksheets[0] if not worksheet else wb_obj[worksheet]
    return ws


def get_sheet_keys(ws):
    for row in ws.values:
        sheet_keys = [x.upper() if isinstance(x, str) else x for x in row]
        break
    return sheet_keys


def shift_col(ws, col_key, target_idx):
    ws.insert_cols(target_idx)
    sheet_keys = get_sheet_keys(ws)
    sheet_key_idx = sheet_keys.index(col_key) + 1  # for openpyxl
    sheet_key_idx_ltr = get_column_letter(sheet_key_idx)
    col_delta = target_idx - sheet_key_idx
    ws.move_range(
        f"{sheet_key_idx_ltr}1:{sheet_key_idx_ltr}{ws.max_row}", rows=0, cols=col_delta
    )
    ws.delete_cols(sheet_key_idx)


def copy_cell_style(ws, style_src_cell, target_cell):
    if style_src_cell.has_style:
        ws[target_cell].font = copy(style_src_cell.font)
        ws[target_cell].border = copy(style_src_cell.border)
        ws[target_cell].fill = copy(style_src_cell.fill)
        ws[target_cell].number_format = copy(style_src_cell.number_format)
        ws[target_cell].protection = copy(style_src_cell.protection)
        ws[target_cell].alignment = copy(style_src_cell.alignment)


def freeze_first_row(worksheet):
    logging.info("Freezing first row of worksheet.")
    worksheet.views.sheetView[0].topLeftCell = "A1"
    worksheet.freeze_panes = "A2"


def create_col(ws, col_key, target_idx, style_src_cell=None):
    ws.insert_cols(target_idx)
    col_header = f"{get_column_letter(target_idx)}1"
    ws[col_header] = col_key.title()
    if style_src_cell:
        copy_cell_style(ws, style_src_cell, col_header)


def sequence_worksheet(ws, col_order, isbn_key):
    sheet_keys = get_sheet_keys(ws)
    for i, key_name in enumerate(col_order):
        order_idx = i + 1  # for openpyxl
        if key_name == "ISBN":
            key_name = isbn_key  # noqa: PLW2901
        if key_name in sheet_keys:
            shift_col(ws, key_name, order_idx)
        else:
            create_col(ws, key_name, order_idx)


def size_sheet_cols(ws, isbn_key):
    dim_holder = DimensionHolder(worksheet=ws)
    sheet_keys = get_sheet_keys(ws)
    for i, key_name in enumerate(sheet_keys):
        col_idx = i + 1  # for openpyxl
        col_idx_ltr = get_column_letter(col_idx)
        width = (
            max(len(str(cell.value)) for cell in ws[col_idx_ltr])
            * CFG["asg"]["spreadsheet"]["sheet_image"]["col_buffer"]
        )
        width = min(width, CFG["asg"]["spreadsheet"]["sheet_image"]["max_col_width"])
        dim_holder[col_idx_ltr] = ColumnDimension(ws, index=col_idx_ltr, width=width)
        if key_name == isbn_key:
            dim_holder[col_idx_ltr] = ColumnDimension(
                ws,
                index=col_idx_ltr,
                width=math.ceil(
                    CFG["asg"]["spreadsheet"]["sheet_image"]["isbn_col_width"]
                    * CFG["asg"]["spreadsheet"]["sheet_image"]["col_buffer"]
                ),
            )
        if key_name == "IMAGE":
            dim_holder[col_idx_ltr] = ColumnDimension(
                ws,
                index=col_idx_ltr,
                width=CFG["asg"]["spreadsheet"]["sheet_image"]["image_col_width"],
            )

    ws.column_dimensions = dim_holder


def insert_image(image_directory, ws, isbn_cell, image_cell):
    namespace = f"{MODULE}.{insert_image.__name__}"
    image_row_height = CFG["asg"]["spreadsheet"]["sheet_image"]["image_row_height"]
    isbn = validate_isbn(isbn_cell.value)
    # Set row height
    row_dim = ws.row_dimensions[image_cell.row]
    row_dim.height = image_row_height

    # Insert image into cell
    filename = f"{isbn}.jpg"
    filepath = os.path.join(image_directory, filename)
    logging.debug(f"{namespace}: Attempting to insert '{filepath}'.")
    if os.path.isfile(filepath):
        img = Image(filepath)
        ws.add_image(img, f"{image_cell.column_letter}{image_cell.row}")
        logging.info(f"{namespace}: Inserted '{filepath}'.")


def sheet_image(vendor_code, workbook, worksheet, image_directory, out):
    namespace = f"{MODULE}.{sheet_image.__name__}"

    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")
    vendr = vendor.Vendor(vendor_code)
    vendr.set_vendor_data()

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")

    col_order = CFG["asg"]["spreadsheet"]["sheet_image"]["col_order"]
    sequence_worksheet(ws, col_order, isbn_key)
    size_sheet_cols(ws, isbn_key)

    # Prepare "IMAGE" column
    sk = get_sheet_keys(ws)
    try:
        img_idx = sk.index("IMAGE") + 1
        img_idx_ltr = get_column_letter(img_idx)
    except ValueError as e:
        logging.error(f"{namespace}: Err finding 'IMAGE' column in sheet '{workbook}'.")
        logging.error("Aborting.")
        raise e
    try:
        isbn_idx = sk.index(isbn_key) + 1
        isbn_idx_ltr = get_column_letter(isbn_idx)
    except ValueError as e:
        logging.error(
            f"{namespace}: Err, no '{isbn_key}' column in sheet '{workbook}'."
        )
        logging.error("Aborting.")
        raise e

    for i in range(1, ws.max_row):
        isbn_cell = ws[f"{isbn_idx_ltr}{i+1}"]
        image_cell = ws[f"{img_idx_ltr}{i+1}"]
        # Format to center content
        image_cell.alignment = Alignment(horizontal="center")
        insert_image(image_directory, ws, isbn_cell, image_cell)

    freeze_first_row(ws)

    # Save workbook
    wb.save(out)


def validate_isbn(isbn):
    # FIXME:  This method is a duplicate from Item.validate_isbn.
    #         This duplication should be removed
    namespace = f"{MODULE}.{validate_isbn.__name__}"
    valid_isbn = ""
    mod_isbn = str(isbn)
    if isbnlib.is_isbn13(mod_isbn) or isbnlib.is_isbn10(mod_isbn):
        valid_isbn = isbnlib.to_isbn13(mod_isbn)
    else:
        # look for formula value
        m = re.search('="(.*)"', mod_isbn)
        if m:
            mod_isbn = m.group(1)
        # look for float value
        mod_isbn = mod_isbn.split(".", 1)[0]
        # look for value with missing zero(s)
        mod_isbn = mod_isbn.zfill(10)
        if isbnlib.is_isbn13(mod_isbn) or isbnlib.is_isbn10(mod_isbn):
            valid_isbn = isbnlib.to_isbn13(mod_isbn)
        else:
            logging.error(f"{namespace}: Err reading isbn '{isbn}'")
    return valid_isbn


def validate_qty(qty):
    namespace = f"{MODULE}.{validate_qty.__name__}"
    try:
        valid_qty = str(int(qty)).strip()
    except Exception as e:
        logging.error(f"{namespace}: Err reading Order qty '{qty}', err: '{e}'")
        valid_qty = None
    return valid_qty


def get_order_items(vendor_code, workbook, worksheet):
    namespace = f"{MODULE}.{get_order_items.__name__}"

    order_items = []
    try:
        order_col = CFG["asg"]["spreadsheet"]["order"]["order_col"].upper()
    except AttributeError:
        logging.error(f"{namespace}: No order column set in config.toml")
        order_col = ""

    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")
    vendr = vendor.Vendor(vendor_code)
    vendr.set_vendor_data()

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")

    # Find Isbn and Order column letters
    row01 = ws[1]
    for cell in row01:
        if cell.value == isbn_key:
            isbn_column_letter = cell.column_letter
        if cell.value.upper() == order_col:
            order_column_letter = cell.column_letter

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter == isbn_column_letter:
                isbn_cell = cell
            if cell.column_letter == order_column_letter:
                order_cell = cell
        # Validate ISBN
        isbn = validate_isbn(isbn_cell.value)
        # Validate Order Qty
        qty = validate_qty(order_cell.value)
        if not isbn and not qty:
            continue
        order_items.append((isbn, qty))

    return order_items


def mkthumbs(image_directory):
    namespace = f"{MODULE}.{mkthumbs.__name__}"

    thumb_width = CFG["asg"]["spreadsheet"]["mkthumbs"]["width"]
    thumb_height = CFG["asg"]["spreadsheet"]["mkthumbs"]["height"]

    here = os.path.dirname(getsourcefile(lambda: 0))
    data = os.path.abspath(os.path.join(here, "data"))
    logo = os.path.join(data, "artemis_logo.png")
    logging.debug(f"{namespace}: Found image for thumbnail background at '{logo}'")
    sub_dir = "thumbnails"
    back = PIL_Image.open(logo)
    thumb_dir = os.path.join(image_directory, sub_dir)
    logging.debug(f"{namespace}: Defining thumbnail directory as '{thumb_dir}'")
    if not os.path.isdir(thumb_dir):
        logging.debug(f"{namespace}: Creating directory '{thumb_dir}'")
        os.mkdir(thumb_dir)
        if os.path.isdir(thumb_dir):
            logging.info(f"{namespace}: Successfully created directory '{thumb_dir}'")
        else:
            logging.error(
                f"{namespace}: Failed to create directory '{thumb_dir}'. Aborting."
            )
            raise Exception
    files = os.listdir(image_directory)
    for f in files:
        # Valid files are JPG or PNG that are not supplemental images.
        image = re.match(r"^.+\.(?:jpg|png)$", f)
        if not image:
            continue
        # Supplemental images have a "-[0-9]+" suffix before the file type.
        # AND a file without that suffix exists int he image_directory.
        suffix = re.match(r"(^.+)-[0-9]+(\.(?:jpg|png))$", f)
        if suffix:
            primary = suffix.group(1) + suffix.group(2)
            primary_path = os.path.join(image_directory, primary)
            if os.path.isfile(primary_path):
                continue
        thumb_file = os.path.join(thumb_dir, f)
        # don't remake thumbnails
        if os.path.isfile(thumb_file):
            continue
        bk = back.copy()
        try:
            file_path = os.path.join(image_directory, f)
            fg = PIL_Image.open(file_path)
        except UnidentifiedImageError:
            logging.error(f"{namespace}: Err reading '{f}', deleting '{file_path}'")
            os.remove(file_path)
            continue
        fg.thumbnail((thumb_width, thumb_height))
        size = (int((bk.size[0] - fg.size[0]) / 2), int((bk.size[1] - fg.size[1]) / 2))
        bk.paste(fg, size)
        logging.debug(f"{namespace}: Attempting to save thumbnail '{thumb_file}'")
        bkn = bk.convert("RGB")
        bkn.save(thumb_file)
        logging.info(f"{namespace}: Successfully created thumbnail '{thumb_file}'")


def get_sheet_data(workbook, worksheet=None):
    namespace = f"{MODULE}.{get_sheet_data.__name__}"
    #########################################################################
    # Try to open sheet_id as an Excel file
    sheet_data = []
    try:
        wb = load_workbook(workbook)
        ws = get_worksheet(wb, worksheet)
        for row in ws.values:
            sheet_data.append(row)
    except (FileNotFoundError, InvalidFileException):
        #########################################################################
        # Google specific stuff
        # authenticate to google sheets
        logging.info(f"{namespace}: Authenticating to google api.")
        creds = app_creds.app_creds()
        sheets_api = build("sheets", "v4", credentials=creds)
        # get sheet data
        if not worksheet:
            sheets = (
                sheets_api.spreadsheets()
                .get(spreadsheetId=workbook)
                .execute()
                .get("sheets", "")
            )
            ws = sheets.pop(0).get("properties", {}).get("title")
        else:
            ws = worksheet
        sheet_data = (
            sheets_api.spreadsheets()
            .values()
            .get(range=ws, spreadsheetId=workbook)
            .execute()
            .get("values")
        )
        #########################################################################
    return sheet_data


def convert_dimensions(dimension):
    """
    Only works for dimensions scraped for Amazon or AmazonUK
    Converts string dimension into width, length, & height.
    Also converts cm to inches

    Example Dimensions:
    "9.84 x 10.63 inches" (width x height)
    >>> convert_dimensions("9.84 x 10.63 inches")
    ('9.84', '', '10.63')

    "5.28 x 0.87 x 7.95 inches" (width x length x height)
    >>> convert_dimensions("5.28 x 0.87 x 7.95 inches")
    ('5.28', '0.87', '7.95')

    "" (no dimension was found)
    >>> convert_dimensions("")
    ('', '', '')

    Returns: [width, length, height]
    """
    width, length, height = "", "", ""
    standard_dimension_count = 3
    needs_unit_conversion = False
    if re.search(r"cm", dimension):  # dimensions in cm
        dimension = re.sub(r"\s*cm\s*", "", dimension)
        needs_unit_conversion = True
    if re.search(r"inches", dimension):  # dimension in inches
        dimension = re.sub(r"\s*inches\s*", "", dimension)  # remove measure unit
    parsed_dimensions = re.split(r" x ", dimension)
    if needs_unit_conversion:  # convert cm to inches
        for idx, dim in enumerate(parsed_dimensions):
            parsed_dimensions[idx] = round((float(dim) / 2.54), 2)
    if len(parsed_dimensions) > 1:
        if len(parsed_dimensions) < standard_dimension_count:  # 2 dimensions, else 3
            width = parsed_dimensions[0]
            height = parsed_dimensions[1]
        else:
            width = parsed_dimensions[0]
            length = parsed_dimensions[1]
            height = parsed_dimensions[2]
    return width, length, height


def get_percentage_discount_prices(price, *discount_percentages):
    price = float(price)
    discount_percent_dict = {}
    for discount in discount_percentages:
        f_disc = float(discount.strip("%")) / 100
        discount_percent_dict[discount] = price - (price * f_disc)
    return discount_percent_dict


def waves_set_pound_price(headers, item, row):
    map_scheme = CFG["asg"]["spreadsheet"]["sheet_waves"]["pound_pricing_map"]
    unmapped_multiplier = CFG["asg"]["spreadsheet"]["sheet_waves"][
        "pound_pricing_unmapped_multipier"
    ]
    rrp = item.data.get("RRP")
    book_format = item.data.get("FORMAT")
    if rrp and "Pound Pricing" in headers:
        try:
            rrp = float(rrp)
        except ValueError:
            return
        pounds, pence = f"{rrp:.2f}".split(".")
        try:
            pound_price = map_scheme[book_format][pounds][pence]
        except KeyError:
            pound_price = rrp * unmapped_multiplier
        pound_price = f"{round(pound_price, 2):.2f}"
        row[headers.index("Pound Pricing")].value = pound_price


def waves_calculate_fields(headers, row, fields):
    """
    fields: A nested dictionary with the following structure:
    First level: Name of column/field to change for a given row.
        Second level:
            map_from:
                Names of columns to access values for a given row.
                These values are used as keys, for indexing into map.
            map:
                Nested dict.
    """
    for field in list(fields.keys()):
        map_from = fields[field]["map_from"]
        field_map = fields[field]["map"]
        if isinstance(map_from, str):
            map_from = [map_from]
        if field in headers:
            current_level = field_map
            for idx, field_key in enumerate(map_from):
                try:
                    field_val = row[headers.index(field_key)].value
                    field_val = "" if field_val is None else field_val
                    if idx == len(map_from) - 1:
                        row[headers.index(field)].value = current_level[field_val]
                    else:
                        current_level = current_level[field_val]
                except KeyError:
                    break
                except ValueError:
                    break


def rename_columns(ws, headers, rename_fields):
    for old_name, new_name in rename_fields.items():
        if old_name in headers:
            col_idx = headers.index(old_name) + 1
            ws.cell(row=1, column=col_idx, value=new_name)


def add_waves_row_data(headers, item, row, gbp_to_usd):
    """
    Adds width, height, length, discounts,
    description, images, and preset fields to a row.
    """
    preset_fields = CFG["asg"]["spreadsheet"]["sheet_waves"]["preset_fields"]
    image_columns = CFG["asg"]["spreadsheet"]["sheet_waves"]["image_columns"]
    discounts = CFG["asg"]["spreadsheet"]["sheet_waves"]["discounted_prices"]
    disc_fstr = CFG["asg"]["spreadsheet"]["sheet_waves"]["discount_text_map"]
    # convert dimension to w, h, l
    width, length, height = convert_dimensions(item.data["DIMENSION"])
    # set new dimension columns
    if "Width" in headers:
        row[headers.index("Width")].value = width
    if "Height" in headers:
        row[headers.index("Height")].value = height
    if "Length" in headers:
        row[headers.index("Length")].value = length
    rrp = item.data.get("RRP")
    if rrp and gbp_to_usd:  # set discounted prices
        discounted_prices = get_percentage_discount_prices(rrp, *discounts)
        for disc, val in discounted_prices.items():
            disc_idx = headers.index(disc_fstr.format(t=disc))
            disc_price = val * gbp_to_usd
            row[disc_idx].value = f"{round(disc_price, 2):.2f}"
    waves_set_pound_price(headers, item, row)
    for idx, img in enumerate(image_columns):  # add image urls
        try:
            row[headers.index(img)].value = item.image_urls[idx]
        except IndexError:
            break
    if "Description" in headers:
        row[headers.index("Description")].value = item.data.get("DESCRIPTION")
    for field, val in preset_fields.items():
        row[headers.index(field)].value = val
    fields = CFG["asg"]["spreadsheet"]["sheet_waves"]["calculate_fields"]
    waves_calculate_fields(headers, row, fields)


def sheet_waves(vendor_code, workbook, worksheet, out, scraped_items_db, gbp_to_usd):
    namespace = f"{MODULE}.{sheet_waves.__name__}"

    addl_data_columns = CFG["asg"]["spreadsheet"]["sheet_waves"]["data_columns"]
    addl_image_columns = CFG["asg"]["spreadsheet"]["sheet_waves"]["image_columns"]
    discounts = CFG["asg"]["spreadsheet"]["sheet_waves"]["discounted_prices"]
    disc_fstr = CFG["asg"]["spreadsheet"]["sheet_waves"]["discount_text_map"]
    preset_fields = CFG["asg"]["spreadsheet"]["sheet_waves"]["preset_fields"]
    for discount in discounts:
        addl_data_columns.append(disc_fstr.format(t=discount))
    for field in preset_fields:
        addl_data_columns.append(field)
    addl_columns = addl_data_columns + addl_image_columns
    # get vendor info from database
    logging.debug(f"{namespace}: Instantiate vendor.")
    vendr = vendor.Vendor(vendor_code)
    vendr.set_vendor_data()

    isbn_key = vendr.isbn_key
    logging.debug(f"{namespace}: Setting ISBN_KEY to '{isbn_key}'.")

    sheet_data = get_sheet_data(workbook, worksheet)

    sheet_keys = [x for x in sheet_data.pop(0) if x]  # filter out None
    items_obj = items.Items(sheet_keys, sheet_data, vendr.isbn_key)
    items_obj.load_scraped_data(scraped_items_db)

    # Load worksheet
    logging.info(f"{namespace}: Workbook is {workbook}")
    wb = load_workbook(workbook)
    ws = get_worksheet(wb, worksheet)
    logging.info(f"{namespace}: Worksheet is {ws.title}")

    # Append columns
    col_insert_idx = ws.max_column + 1
    ws.insert_cols(col_insert_idx, len(addl_columns))
    i = 1
    for col in addl_columns:
        col_idx = col_insert_idx + i
        ws.cell(row=1, column=col_idx, value=col)
        i = i + 1

    # Find ISBN column
    row01 = ws[1]
    isbn_idx = None
    for cell in row01:
        if isinstance(cell.value, str) and cell.value.upper() == isbn_key.upper():
            isbn_idx = cell.column - 1
            break
    if isbn_idx is None:
        logging.error(f"{namespace}: Err no isbn column in spreadsheet")
        raise Exception
    # get column headers
    headers = [cell.value for cell in row01]
    # Insert data into cells
    for row in ws.iter_rows(min_row=2):
        # get isbn cell
        isbn = str(row[isbn_idx].value)
        # find items_obj matching isbn
        item = items_obj.find_item(isbn)
        if item:
            add_waves_row_data(headers, item, row, gbp_to_usd)
    rename_fields = CFG["asg"]["spreadsheet"]["sheet_waves"]["rename_fields"]
    rename_columns(ws, headers, rename_fields)
    # Save workbook
    wb.save(out)
